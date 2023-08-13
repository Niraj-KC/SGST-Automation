import os
import openpyxl


eligibleSGSTAdjustedAgainstOtherGoods_dict = {}

filepath = r'TestData/GST CLAIM SUMMARY.xlsx'

wb = openpyxl.load_workbook(filepath)
ws = wb['Sheet1']


def wordfinder(searchString):
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if searchString == ws.cell(i, j).value:
                # print("found")
                # print(ws.cell(i, j))
                return i, j

def getEligibleSGSTAdjustedAgainstOtherGoods():

    row, col = wordfinder("Eligible SGST Adjusted againnst in eligible goods")
    row_RCM, col_RCM = wordfinder("Eligible SGST RCM  input ")
    row += 1
    for i in range(row, ws.max_row+1):
        print(f'{ws.cell(i, 2).value}    ==    {ws.cell(i, col).value} |||  {ws.cell(i, col_RCM)}')


getEligibleSGSTAdjustedAgainstOtherGoods()



